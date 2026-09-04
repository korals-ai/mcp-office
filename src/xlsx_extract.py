"""Read an Excel workbook's real cell values via openpyxl.

Companion to ``office_convert.py``/``pdf_text.py``: those give the agent a
document's *rendered* form (another file format, or a PDF's flat text);
this module gives it the workbook's actual **structured** data — per-sheet
rows of typed cell values (numbers, strings, dates, formula RESULTS via
``data_only=True``), not a flattened text/CSV dump that would lose column
alignment across merged cells or blank columns.

Before this existed, the only way for the agent to read an xlsx's cells was
raw Bash + an inline openpyxl script (see the metrobit tender-reconciliation
chat, 2026-09-04) — this makes that a first-class, tested tool instead of an
ad-hoc shell one-liner, matching how ``pdf_extract_text`` replaced raw
``pdftotext`` calls.

Importable contract:

    from pathlib import Path
    from src.xlsx_extract import extract_cells, XlsxExtractError

    sheets = extract_cells(Path("/home/agent/costsheet.xlsx"))
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Any

logger = logging.getLogger(__name__)

# Same rationale as pdf_text.MAX_TEXT_CHARS: this return value goes straight
# into the agent's tool_result / context, and an oversized message previously
# killed a whole SDK session (docs/incidents/2026-08-04-workspace-sdk-1mib-
# message-buffer.md). A tender costsheet is normally tens to low hundreds of
# rows; this is a generous ceiling, not a target.
MAX_CELLS = 50_000


class XlsxExtractError(RuntimeError):
    """Reading the workbook failed: missing/corrupt source, unknown sheet
    name, or the cell count exceeds ``MAX_CELLS``."""


def _jsonable(value: object) -> Any:
    """Coerce an openpyxl cell value to something JSON-serializable over the
    MCP wire. openpyxl already returns plain ``int``/``float``/``str``/``bool``/
    ``None`` for ordinary cells; only date/time types need converting."""
    import datetime

    if isinstance(value, datetime.datetime | datetime.date | datetime.time):
        return value.isoformat()
    return value


def extract_cells(
    src: Path,
    *,
    sheet: str | None = None,
) -> dict[str, dict[str, object]]:
    """Read every sheet's cell values (or just ``sheet``, if given) from the
    ``.xlsx`` at ``src``.

    Returns ``{sheet_title: {"dimensions": "A1:D10", "rows": [[...], ...]}}``
    — one entry per sheet, in workbook order. ``rows`` is every row in the
    sheet's used range (``ws.dimensions``), each a list of cell values in
    column order; a wholly blank row is still included so row numbers line up
    with the source spreadsheet. Formulas are read as their last-computed
    result (``data_only=True``), not the formula text — this tool is for
    reconciling values, not auditing formulas.

    Raises ``XlsxExtractError`` for a missing/non-file/corrupt source, an
    unknown ``sheet`` name, or a workbook whose cell count exceeds
    ``MAX_CELLS``.
    """
    if not src.exists():
        raise XlsxExtractError(f"source does not exist: {src}")
    if not src.is_file():
        raise XlsxExtractError(f"source is not a file: {src}")

    import zipfile

    from openpyxl import load_workbook
    from openpyxl.utils.exceptions import InvalidFileException

    try:
        # Not read_only: the read-only worksheet drops `.dimensions`, and
        # these workbooks (tender costsheets, not million-row datasets) are
        # small enough that the fully-loaded model's extra memory is a
        # non-issue.
        wb = load_workbook(src, data_only=True)
    except (InvalidFileException, KeyError, zipfile.BadZipFile, OSError) as exc:
        # openpyxl raises a bare KeyError for some malformed OOXML zips.
        raise XlsxExtractError(f"could not open {src.name} as an xlsx: {exc}") from exc

    try:
        titles = wb.sheetnames
        if sheet is not None:
            if sheet not in titles:
                raise XlsxExtractError(
                    f"sheet {sheet!r} not found in {src.name} — available: {titles}"
                )
            titles = [sheet]

        result: dict[str, dict[str, object]] = {}
        total_cells = 0
        for title in titles:
            ws = wb[title]
            rows = [[_jsonable(cell) for cell in row] for row in ws.iter_rows(values_only=True)]
            total_cells += sum(len(row) for row in rows)
            if total_cells > MAX_CELLS:
                raise XlsxExtractError(
                    f"{src.name} exceeds {MAX_CELLS} cells read so far — too large for "
                    "this tool; ask the user to split it or narrow the 'sheet' argument"
                )
            result[title] = {"dimensions": ws.dimensions, "rows": rows}
        return result
    finally:
        wb.close()
